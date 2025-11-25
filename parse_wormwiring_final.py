#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extract neuron connection data from WormWiring Excel files and build graph structure
- Only consider hermaphrodite
- Chemical synapses and gap junction (asymmetric)
- Only consider I (interneuron), S (sensory), M (motor) neurons

Worksheet structure:
- Row 1: Type grouping labels (column header position)
- Row 3: Neuron names (column headers, starting from column 4)
- Column 1: PHARYNX and other non-neurons
- Column 3: Neuron names (row headers, starting from row 4)
- Data area: starting from row 4, column 4
"""
import openpyxl
import networkx as nx
from pathlib import Path
import pickle
import pandas as pd
import re

def find_type_boundaries(ws):
    """Find type grouping boundaries (columns and rows), including excluded type boundaries"""
    # Find type grouping column indices in row 1
    type_cols = {}  # {column_index: type_name}
    exclude_boundaries_col = []  # Column boundaries for excluded types
    
    for j in range(1, ws.max_column + 1):
        val = ws.cell(1, j).value
        if val is not None:
            val_str = str(val).strip().upper()
            # Record boundaries for excluded types
            if any(exclude in val_str for exclude in ['BODYWALL MUSCLES', 'OTHER END ORGANS', 'SEX-SPECIFIC CELLS', 'SEX SPECIFIC', 'PHARYNX']):
                exclude_boundaries_col.append(j)
            elif 'SENSORY NEURONS' in val_str:
                type_cols[j] = 'S'
            elif 'INTERNEURONS' in val_str or ('INTER' in val_str and 'NEURON' in val_str):
                type_cols[j] = 'I'
            elif 'MOTOR NEURONS' in val_str:
                type_cols[j] = 'M'
    
    # Find type grouping row indices in column 1
    type_rows = {}  # {row_index: type_name}
    exclude_boundaries_row = []  # Row boundaries for excluded types
    
    for i in range(1, ws.max_row + 1):
        val = ws.cell(i, 1).value
        if val is not None:
            val_str = str(val).strip().upper()
            # Record boundaries for excluded types
            if any(exclude in val_str for exclude in ['BODYWALL MUSCLES', 'OTHER END ORGANS', 'SEX-SPECIFIC CELLS', 'SEX SPECIFIC', 'PHARYNX']):
                exclude_boundaries_row.append(i)
            elif 'SENSORY NEURONS' in val_str:
                type_rows[i] = 'S'
            elif 'INTERNEURONS' in val_str or ('INTER' in val_str and 'NEURON' in val_str):
                type_rows[i] = 'I'
            elif 'MOTOR NEURONS' in val_str:
                type_rows[i] = 'M'
    
    return type_cols, type_rows, exclude_boundaries_col, exclude_boundaries_row

def col_letter_to_num(col_letter):
    """Convert Excel column letter (e.g., 'BB', 'LM') to column number"""
    result = 0
    for char in col_letter:
        result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result

def extract_neurons(ws):
    """Extract neuron list and their types - only extract three neuron types (SENSORY, INTERNEURON, MOTOR)"""
    # Find type grouping boundaries
    type_cols, type_rows, exclude_cols, exclude_rows = find_type_boundaries(ws)
    
    # Extract column headers (row 3) - target neurons
    neuron_cols = {}  # {neuron_name: (column_index, type)}
    col_boundaries = sorted(type_cols.items())
    
    for j in range(4, ws.max_column + 1):
        # Check if within excluded type range (BODYWALL MUSCLES, etc.)
        in_exclude = False
        if exclude_cols:
            for exc_col in sorted(exclude_cols):
                if j >= exc_col:
                    # Find next neuron type grouping
                    next_neuron_col = None
                    for col_idx, _ in col_boundaries:
                        if col_idx > exc_col:
                            next_neuron_col = col_idx
                            break
                    # If current column is between excluded type and next neuron type, exclude it
                    if next_neuron_col is None or j < next_neuron_col:
                        in_exclude = True
                        break
        
        if in_exclude:
            continue
        
        # Determine which type the current column belongs to
        neuron_type = None
        for idx, (col_idx, ntype) in enumerate(col_boundaries):
            if j >= col_idx:
                neuron_type = ntype
            if idx + 1 < len(col_boundaries) and j < col_boundaries[idx + 1][0]:
                break
        
        if neuron_type is None:
            continue
        
        val = ws.cell(3, j).value  # Row 3 contains column headers
        if val is not None:
            neuron_name = str(val).strip()
            if neuron_name and len(neuron_name) < 20:
                neuron_cols[neuron_name] = (j, neuron_type)
    
    # Extract row headers (column 3) - source neurons
    neuron_rows = {}  # {neuron_name: (row_index, type)}
    row_boundaries = sorted(type_rows.items())
    
    for i in range(4, ws.max_row + 1):
        # Check row header, exclude non-neuron types
        val_header = ws.cell(i, 1).value
        if val_header is not None:
            header_str = str(val_header).strip().upper()
            if any(exclude in header_str for exclude in ['BODYWALL MUSCLES', 'OTHER END ORGANS', 'SEX-SPECIFIC', 'PHARYNX']):
                continue
        
        # Determine which type the current row belongs to
        neuron_type = None
        for idx, (row_idx, ntype) in enumerate(row_boundaries):
            if i >= row_idx:
                neuron_type = ntype
            if idx + 1 < len(row_boundaries) and i < row_boundaries[idx + 1][0]:
                break
        
        if neuron_type is None:
            continue
        
        val = ws.cell(i, 3).value  # Column 3 contains row headers
        if val is not None:
            neuron_name = str(val).strip()
            if neuron_name and len(neuron_name) < 20:
                # Only include nodes that exist in neuron_cols (ensures they are one of the three neuron types)
                if neuron_name in neuron_cols:
                    neuron_rows[neuron_name] = (i, neuron_type)
    
    return neuron_rows, neuron_cols

def build_graph(excel_path):
    """Build graph"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Find worksheets
    ws_chemical = None
    ws_gap = None
    for name in wb.sheetnames:
        if 'hermaphrodite' in name.lower() and 'chemical' in name.lower():
            ws_chemical = wb[name]
        if 'hermaphrodite' in name.lower() and 'gap' in name.lower() and 'asymmetric' in name.lower():
            ws_gap = wb[name]
    
    if ws_chemical is None or ws_gap is None:
        raise ValueError(f"Required worksheets not found")
    
    print("=" * 80)
    print("Parsing WormWiring Excel file")
    print("=" * 80)
    print(f"Using worksheets: {ws_chemical.title}, {ws_gap.title}")
    
    # Extract neuron list
    print("\n1. Extracting neuron list...")
    neurons_rows_chem, neurons_cols_chem = extract_neurons(ws_chemical)
    neurons_rows_gap, neurons_cols_gap = extract_neurons(ws_gap)
    
    # Merge neuron lists (only keep I/S/M types)
    all_neurons = {}
    for name, (row_idx, ntype) in neurons_rows_chem.items():
        if ntype in ['S', 'I', 'M']:
            all_neurons[name] = ntype
    for name, (col_idx, ntype) in neurons_cols_chem.items():
        if ntype in ['S', 'I', 'M']:
            if name not in all_neurons:
                all_neurons[name] = ntype
    
    print(f"  Found {len(all_neurons)} neurons (I/S/M types)")
    print(f"  S={sum(1 for t in all_neurons.values() if t=='S')}, "
          f"I={sum(1 for t in all_neurons.values() if t=='I')}, "
          f"M={sum(1 for t in all_neurons.values() if t=='M')}")
    
    # Create graph
    G = nx.DiGraph()
    for name, ntype in all_neurons.items():
        G.add_node(name, neuron_type=ntype)
    
    # Add edges from chemical synapse table
    print("\n2. Adding chemical synapse edges...")
    chemical_count = 0
    for source_name, (source_row, _) in neurons_rows_chem.items():
        if source_name not in all_neurons:
            continue
        for target_name, (target_col, _) in neurons_cols_chem.items():
            if target_name not in all_neurons:
                continue
            
            weight = ws_chemical.cell(source_row, target_col).value
            if weight is not None:
                try:
                    w = float(weight)
                    if w > 0:
                        G.add_edge(source_name, target_name,
                                 chemical_weight=w, edge_type='chemical')
                        chemical_count += 1
                except (ValueError, TypeError):
                    pass
    
    print(f"  Added {chemical_count} chemical synapse edges")
    
    # Add edges from gap junction (asymmetric) table
    print("\n3. Adding gap junction edges...")
    gap_count = 0
    both_count = 0
    for source_name, (source_row, _) in neurons_rows_gap.items():
        if source_name not in all_neurons:
            continue
        for target_name, (target_col, _) in neurons_cols_gap.items():
            if target_name not in all_neurons:
                continue
            
            weight = ws_gap.cell(source_row, target_col).value
            if weight is not None:
                try:
                    w = float(weight)
                    if w > 0:
                        if G.has_edge(source_name, target_name):
                            G[source_name][target_name]['gap_weight'] = w
                            G[source_name][target_name]['edge_type'] = 'both'
                            both_count += 1
                        else:
                            G.add_edge(source_name, target_name,
                                     gap_weight=w, edge_type='gap')
                            gap_count += 1
                except (ValueError, TypeError):
                    pass
    
    print(f"  Added {gap_count} gap junction edges")
    print(f"  {both_count} edges have both chemical and gap connections")
    
    return G, all_neurons

def main():
    excel_path = Path("data/raw/wormwiring_SI5_connectome_adjacency_corrected_2020.xlsx")
    
    if not excel_path.exists():
        print(f"Error: File does not exist {excel_path}")
        return
    
    G, neurons = build_graph(excel_path)
    
    # Statistics
    print("\n" + "=" * 80)
    print("Graph Statistics")
    print("=" * 80)
    print(f"Number of nodes: {G.number_of_nodes()}")
    print(f"Number of edges: {G.number_of_edges()}")
    
    type_counts = {'S': 0, 'I': 0, 'M': 0}
    for node, data in G.nodes(data=True):
        ntype = data.get('neuron_type', '')
        if ntype in type_counts:
            type_counts[ntype] += 1
    
    print(f"\nNode type distribution:")
    for ntype, count in type_counts.items():
        print(f"  {ntype}: {count}")
    
    # Count edge types
    edge_type_counts = {'chemical': 0, 'gap': 0, 'both': 0}
    for u, v, d in G.edges(data=True):
        etype = d.get('edge_type', 'unknown')
        if etype in edge_type_counts:
            edge_type_counts[etype] += 1
    
    print(f"\nEdge type distribution:")
    print(f"  Chemical only: {edge_type_counts['chemical']}")
    print(f"  Gap junction only: {edge_type_counts['gap']}")
    print(f"  Both chemical and gap: {edge_type_counts['both']}")
    
    # Save
    output_dir = Path("data/processed")
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Save graph
    graph_path = output_dir / "graph_wormwiring_hermaphrodite_ISM.pickle"
    with open(graph_path, 'wb') as f:
        pickle.dump(G, f)
    print(f"\n✓ Graph saved: {graph_path}")
    
    # Save CSV
    edges_data = []
    for u, v, d in G.edges(data=True):
        edges_data.append({
            'source': u,
            'target': v,
            'source_type': G.nodes[u]['neuron_type'],
            'target_type': G.nodes[v]['neuron_type'],
            'chemical_weight': d.get('chemical_weight', 0),
            'gap_weight': d.get('gap_weight', 0),
            'edge_type': d.get('edge_type', 'unknown')
        })
    
    df_edges = pd.DataFrame(edges_data)
    edges_csv = output_dir / "edges_wormwiring_ISM.csv"
    df_edges.to_csv(edges_csv, index=False)
    print(f"✓ Edge list saved: {edges_csv}")
    
    nodes_data = [{'node_id': n, 'neuron_type': d['neuron_type']} 
                  for n, d in G.nodes(data=True)]
    df_nodes = pd.DataFrame(nodes_data)
    nodes_csv = output_dir / "nodes_wormwiring_ISM.csv"
    df_nodes.to_csv(nodes_csv, index=False)
    print(f"✓ Node list saved: {nodes_csv}")
    
    # Display examples
    print("\n" + "=" * 80)
    print("Example nodes (first 15):")
    print("=" * 80)
    for i, (node, data) in enumerate(list(G.nodes(data=True))[:15]):
        print(f"  {node}: {data['neuron_type']}")
    
    print("\nExample edges (first 10):")
    print("=" * 80)
    for i, (u, v, d) in enumerate(list(G.edges(data=True))[:10]):
        chem = d.get('chemical_weight', 0)
        gap = d.get('gap_weight', 0)
        print(f"  {u} -> {v}: chemical={chem}, gap={gap}, type={d.get('edge_type')}")

if __name__ == "__main__":
    main()

